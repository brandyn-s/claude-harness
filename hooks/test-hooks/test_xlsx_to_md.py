"""Tests for xlsx-to-md.py (PreToolUse:Read)."""
import os

from conftest import run_hook

HOOK = "xlsx-to-md.py"


def _make_xlsx(path):
    """Build a two-sheet xlsx fixture using openpyxl."""
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove the default sheet and create named ones explicitly
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws1 = wb.create_sheet(title="Findings")
    ws1.append(["ID", "Severity", "Status"])
    ws1.append(["RULE-000000", "high", "open"])
    ws1.append(["RULE-000000", "medium", "not_a_finding"])
    ws2 = wb.create_sheet(title="Meta")
    ws2.append(["Key", "Value"])
    ws2.append(["System", "ExampleOne"])
    wb.save(str(path))


def test_non_xlsx_passthrough():
    rc, _out, err = run_hook(HOOK, {
        "tool_name": "Read",
        "tool_input": {"file_path": "$HOME/Documents/notes.txt"},
    })
    assert rc == 0
    assert err.strip() == ""


def test_non_read_tool_passthrough():
    rc, _out, _err = run_hook(HOOK, {
        "tool_name": "Bash",
        "tool_input": {"command": "start sheet.xlsx"},
    })
    assert rc == 0


def test_nonexistent_file_passthrough():
    rc, _out, _err = run_hook(HOOK, {
        "tool_name": "Read",
        "tool_input": {"file_path": "C:/nonexistent/sheet.xlsx"},
    })
    assert rc == 0


def test_conversion_blocks_with_redirect(tmp_path):
    src = tmp_path / "sheet.xlsx"
    _make_xlsx(src)

    rc, _out, err = run_hook(HOOK, {
        "tool_name": "Read",
        "tool_input": {"file_path": str(src)},
    }, timeout=60)

    assert rc == 2, f"Expected block (2), got {rc}. stderr={err}"
    assert "Read this file instead:" in err

    md_path = err.split("Read this file instead:")[-1].strip()
    assert os.path.isfile(md_path)
    content = open(md_path, "r", encoding="utf-8").read()
    assert content.startswith("<!-- xlsx-to-md-hook -->")
    assert "## Sheet: Findings" in content
    assert "## Sheet: Meta" in content
    assert "RULE-000000" in content
    assert "ExampleOne" in content

    try:
        os.remove(md_path)
    except OSError:
        pass
