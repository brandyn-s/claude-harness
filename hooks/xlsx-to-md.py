"""PreToolUse hook for Read: Auto-convert .xlsx workbooks to Markdown.

Intercepts Read tool calls on .xlsx files, converts each sheet into a markdown
table, caches in ~/Documents/converted-xlsx/, and blocks the Read with a redirect
to the .md file.

Converter-hook pattern: sentinel header, freshness check, subprocess isolation,
paths via argv. Requires openpyxl in the target Python environment.
"""

import json
import os
import subprocess
import sys
from pathlib import Path

CREATE_NO_WINDOW = 0x08000000 if sys.platform == "win32" else 0

_HOME = str(Path.home())


def _resolve_python_exe():
    """Pick the Python interpreter for the conversion subprocess. See
    cklb-to-md.py docstring of the same name for the priority rationale
    (env override → author's local path → sys.executable). Resolves the
    Windows-CI WinError 2 class. Conversion subprocess requires
    openpyxl in the chosen interpreter's site-packages — both the
    author's Python 3.12 and the CI runner's Python (via
    requirements-dev.txt) have it."""
    explicit = os.environ.get("PYTHON_EXE")
    if explicit:
        return explicit
    if sys.platform == "win32":
        local = Path(_HOME, "AppData/Local/Programs/Python/Python312/python.exe")
        if local.exists():
            return str(local)
    return sys.executable


PYTHON_312 = _resolve_python_exe()

SENTINEL = "<!-- xlsx-to-md-hook -->"
CACHE_DIR = str(Path(_HOME, "Documents/converted-xlsx"))
DOCUMENTS_ROOT = str(Path(_HOME, "Documents"))

CONVERT_SCRIPT = r"""
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("NO_OPENPYXL")
    sys.exit(2)

SENTINEL = "<!-- xlsx-to-md-hook -->"
CELL_MAX = 500

src_path = sys.argv[1]
md_path = sys.argv[2]

os.makedirs(os.path.dirname(md_path), exist_ok=True)

try:
    wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
except Exception as e:
    print(f"LOAD_ERROR: {e}")
    sys.exit(1)

def cell_str(v):
    if v is None:
        return ""
    s = str(v).replace("|", "\\|").replace("\r", " ").replace("\n", " ").strip()
    if len(s) > CELL_MAX:
        s = s[:CELL_MAX].rstrip() + " ..."
    return s

lines = []
lines.append(SENTINEL)
lines.append(f"<!-- source: {src_path} -->")
lines.append("")
lines.append(f"# Workbook: {os.path.basename(src_path)}")
lines.append("")
lines.append(f"Sheets: {len(wb.sheetnames)}")
lines.append("")
for name in wb.sheetnames:
    lines.append(f"- {name}")
lines.append("")

for name in wb.sheetnames:
    ws = wb[name]
    lines.append(f"## Sheet: {name}")
    lines.append("")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        lines.append("_(empty)_")
        lines.append("")
        continue

    # Determine column count from the first non-empty row or header
    header = list(first) if first else []
    # Trim trailing all-None columns from header
    while header and header[-1] is None:
        header.pop()
    ncols = max(1, len(header))
    if not any(c is not None and str(c).strip() for c in header):
        # Blank header row — generate column labels
        header = [f"col{i+1}" for i in range(ncols)]

    # Normalize header length
    header = [cell_str(h) if h is not None else f"col{i+1}" for i, h in enumerate(header)]

    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "|".join(["---"] * ncols) + "|")

    row_count = 0
    for row in rows_iter:
        if row is None:
            continue
        # Skip fully empty rows
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        vals = list(row)[:ncols]
        while len(vals) < ncols:
            vals.append("")
        lines.append("| " + " | ".join(cell_str(v) for v in vals) + " |")
        row_count += 1
        if row_count >= 5000:
            lines.append(f"| _... truncated at 5000 rows_ |" + " |" * (ncols - 1))
            break
    lines.append("")

wb.close()

with open(md_path, "w", encoding="utf-8") as fh:
    fh.write("\n".join(lines))

print("OK")
"""


def _get_md_path(src_path):
    norm_src = os.path.normpath(src_path).replace("\\", "/")
    norm_root = os.path.normpath(DOCUMENTS_ROOT).replace("\\", "/")
    if norm_src.lower().startswith(norm_root.lower() + "/"):
        rel = norm_src[len(norm_root) + 1:]
    else:
        rel = os.path.basename(norm_src)
    rel_md = os.path.splitext(rel)[0] + ".md"
    return os.path.join(CACHE_DIR, rel_md).replace("\\", "/")


def block(reason):
    print(reason, file=sys.stderr)
    sys.exit(2)


def warn(message):
    print(message, file=sys.stderr)
    sys.exit(0)


def main():
    try:
        hook_input = json.loads(sys.stdin.read())
    except (json.JSONDecodeError, EOFError):
        sys.exit(0)

    if hook_input.get("tool_name") != "Read":
        sys.exit(0)

    file_path = hook_input.get("tool_input", {}).get("file_path", "")
    if not file_path.lower().endswith(".xlsx"):
        sys.exit(0)

    if not os.path.isfile(file_path):
        sys.exit(0)

    md_path = _get_md_path(file_path)

    if os.path.isfile(md_path):
        try:
            with open(md_path, "r", encoding="utf-8") as f:
                first_line = f.readline().strip()
        except (OSError, UnicodeDecodeError):
            first_line = ""

        if first_line == SENTINEL:
            src_mtime = os.path.getmtime(file_path)
            md_mtime = os.path.getmtime(md_path)
            md_size = os.path.getsize(md_path)
            if md_mtime >= src_mtime and md_size > len(SENTINEL) + 10:
                block(
                    f"XLSX auto-converted to Markdown (cached, {md_size:,} bytes). "
                    f"Read this file instead: {md_path}"
                )
        else:
            warn(
                f"Existing .md file found at {md_path} (not auto-generated). "
                "Skipping XLSX conversion to avoid overwriting. "
                "Delete or rename the .md file to enable auto-conversion."
            )

    try:
        src_size = os.path.getsize(file_path)
        src_mb = src_size / 1024 / 1024

        proc = subprocess.run(
            [PYTHON_312, "-c", CONVERT_SCRIPT, file_path, md_path],
            capture_output=True,
            text=True,
            timeout=295,
            encoding="utf-8",
            errors="replace",
            creationflags=CREATE_NO_WINDOW,
        )
        if proc.returncode == 0 and "OK" in proc.stdout:
            md_size = os.path.getsize(md_path) if os.path.isfile(md_path) else 0
            block(
                f"XLSX auto-converted to Markdown ({md_size:,} bytes from {src_mb:.2f}MB .xlsx). "
                f"Read this file instead: {md_path}"
            )

        if os.path.isfile(md_path):
            try:
                with open(md_path, "r", encoding="utf-8") as f:
                    if f.readline().strip() == SENTINEL:
                        os.remove(md_path)
            except OSError:
                pass

        out_combined = (proc.stdout or "") + (proc.stderr or "")
        if "NO_OPENPYXL" in out_combined:
            warn(
                "XLSX conversion unavailable: openpyxl not installed in target Python. "
                "Run: pip install openpyxl. Falling back to native Read."
            )
        if proc.stderr:
            err_msg = proc.stderr.strip()[:200]
            warn(
                f"XLSX-to-Markdown conversion failed: {err_msg}. "
                "Falling back to native Read."
            )

    except subprocess.TimeoutExpired:
        if os.path.isfile(md_path):
            try:
                with open(md_path, "r", encoding="utf-8") as f:
                    if f.readline().strip() == SENTINEL:
                        os.remove(md_path)
            except OSError:
                pass
        warn(
            f"XLSX conversion timed out (>295s, {src_mb:.2f}MB). "
            "Falling back to native Read."
        )
    except (FileNotFoundError, OSError) as e:
        warn(f"XLSX conversion error: {e}. Falling back to native Read.")

    sys.exit(0)


if __name__ == "__main__":
    # crash-safety: wrap main() so an unhandled exception exits 0
    try:
        main()
    except SystemExit:
        raise
    except Exception:
        sys.exit(0)